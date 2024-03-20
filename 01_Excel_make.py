# 패키지 설치
# PS C:\Users\dell\python> pip install openpyxl
# PS C:\Users\dell\python> pip install pandas
# PS C:\Users\dell\python> pip install xlwings

import openpyxl   # 공식문서 https://openpyxl.readthedocs.io/en/stable/

wb = openpyxl.Workbook()

ws = wb.create_sheet('오징어게임')

ws['A1'] = '참가번호'
ws['B1'] = '성명'

ws['A2'] = 1
ws['B2'] = '오일남'

wb.save(r'C:\Users\dell\python\Excel_make.xlsx')

