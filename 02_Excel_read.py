# 패키지 설치
# PS C:\Users\dell\python> pip install openpyxl
# PS C:\Users\dell\python> pip install pandas
# PS C:\Users\dell\python> pip install xlwings

import openpyxl    # 공식문서 https://openpyxl.readthedocs.io/en/stable/

fpath = r'C:\Users\dell\python\Excel_make.xlsx'

wb = openpyxl.load_workbook(fpath)

ws = wb['오징어게임']

ws['A3'] = 456
ws['B3'] = '성기훈'

wb.save(fpath)

